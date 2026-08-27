#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Aqyl Energy — CEO dashboard auto-updater (v2).

Запуск:
    python3 update.py            # собрать + запушить в GitHub
    python3 update.py --dry-run  # собрать index.html, но НЕ пушить

Источники (по ТЗ «доработки по дашборду»):
  1. Установки (MMS)      — Google Sheets «Ежедневный отчёт по установке»
                            (все листы: Ввод данных, ШПФ, ТПФ, КПФ)
  2. Склад                — Google Sheets «счётчики отгрузка на склад факт»
  3. Активировано + e-Qural total — PostgreSQL e-Qural (read-only)
  4. План                 — хардкод-конфиг ниже

Требования:
    pip install --break-system-packages psycopg2-binary pandas requests openpyxl
    VPN GlobalProtect поднят (для БД).
"""

import sys
import io
import re
import json
import subprocess
from datetime import date, datetime, timedelta

import pandas as pd
import requests

# ──────────────────────────────────────────────────────────────
# КОНФИГ
# ──────────────────────────────────────────────────────────────

DB = {
    "host": "10.20.43.58",
    "port": 5432,
    "dbname": "mds",
    "user": "bts_digital_ro",
    "password": "da%bDK#!M,P4On5",
}

# БД MMS (MySQL) — источник «Активировано» и «К установке»
DB_MMS = {
    "host": "10.20.38.49",
    "port": 3306,
    "database": "device_life",
    "user": "device_life",
    "password": "Device!2026@Halifel#%",
}

# Коды статусов устройства в dl_device.device_status
MMS_STATUS = {
    "stock": "1",       # На складе
    "toInstall": "2",   # К установке
    "installed": "3",   # Установлен (= активировано в базе)
    "removed": "4",     # Демонтирован
    "returned": "5",    # Возврат на завод
    "checked": "6",     # Поверено
}
# region_code (KATO) → регион
MMS_REGION = {"79": "shymkent", "61": "turkestan", "43": "kyzylorda"}

# e-Qural API — количество активированных приборов
EQURAL_API_URL = ("https://equral.ktga.kz/api/metering-system/"
                  "MeteringDeviceService/GetActivatedCount")
EQURAL_API_KEY = "siUJOyOFK3ownp5JqdxgqRq7ICZRwwPybCpKkMbxnpFGOqpZZQQpw6X2xiVUeipD"
EQURAL_REGION_ID = {
    "shymkent":  "58476cad-2855-46d9-a53d-fc198c49831e",
    "turkestan": "c66c2de2-68ca-464d-aa65-c6a96b67360e",
    "kyzylorda": "9943fdcc-0e48-48f3-99ef-b105aa162482",
}

# Google Sheets — экспорт всей книги в xlsx (тянет ВСЕ листы разом)
SHEET_INSTALL_ID = "1nrzWmjIK2JFCKT3RTxt5lIj0LupGvF60"   # Ежедневный отчёт по установке
SHEET_WAREHOUSE_ID = "10zFs4L6llvm39kOwStzCvZpeg2SY5g6o"  # Склад
SHEET_WAREHOUSE_GID = "1245358309"

INSTALL_XLSX_URL = (
    f"https://docs.google.com/spreadsheets/d/{SHEET_INSTALL_ID}/export?format=xlsx"
)
WAREHOUSE_CSV_URL = (
    f"https://docs.google.com/spreadsheets/d/{SHEET_WAREHOUSE_ID}"
    f"/export?format=csv&gid={SHEET_WAREHOUSE_GID}"
)

REPO_DIR = "~/Documents/dapper-hamster-8e138c"
INDEX_FILE = "index.html"
TEMPLATE_FILE = "template.html"   # standalone-дизайн с window.__DASH_DATA

PROJECT_START = date(2026, 4, 22)
YEAR_END = date(2026, 12, 31)

PLAN_2026 = 199794
PHASE1_PLAN = 454700
TOTAL_PROGRAM_PLAN = 1352154
PLAN_TO_AUG = 124875   # накопительный план к августу

MONTHLY_PLAN = {
    4: 24975, 5: 24975, 6: 24975, 7: 24975, 8: 24975, 9: 24975,
    10: 15929, 11: 15929, 12: 18086,
}

REGION_NAMES = {"43": "Кызылординская область",
                "61": "Туркестанская область",
                "79": "город Шымкент"}

REGIONAL_PLANS = {
    "Кызылординская область": {"plan2026": 43328, "plan2027": 55000, "phasePlan": 98328},
    "Туркестанская область":  {"plan2026": 125053, "plan2027": 158400, "phasePlan": 283453},
    "город Шымкент":          {"plan2026": 31413, "plan2027": 42300, "phasePlan": 73713},
}

CAPACITY = {"turkestan": 89500, "shymkent": 89500, "kyzylorda": 75000}

ANNUAL_SCHEDULE = [
    {"year": "2026", "phase1": 199794, "phase2": 0,      "total": 199794},
    {"year": "2027", "phase1": 255534, "phase2": 158387, "total": 413921},
    {"year": "2028", "phase1": 0,      "phase2": 365694, "total": 365694},
    {"year": "2029", "phase1": 0,      "phase2": 372745, "total": 372745},
]

# Исторические daily апрель–июль (из IoT платформы, до перехода на MMS-отчёт).
# Формат: {"date","shymkent","turkestan","kyzylorda"}
HISTORICAL_DAILY = json.loads(r'''[{"date":"2026-04-22","shymkent":0,"turkestan":0,"kyzylorda":1},{"date":"2026-04-23","shymkent":0,"turkestan":5,"kyzylorda":6},{"date":"2026-04-24","shymkent":0,"turkestan":9,"kyzylorda":11},{"date":"2026-04-27","shymkent":0,"turkestan":0,"kyzylorda":9},{"date":"2026-04-28","shymkent":0,"turkestan":0,"kyzylorda":19},{"date":"2026-04-29","shymkent":0,"turkestan":31,"kyzylorda":19},{"date":"2026-04-30","shymkent":0,"turkestan":54,"kyzylorda":12},{"date":"2026-05-04","shymkent":30,"turkestan":74,"kyzylorda":15},{"date":"2026-05-05","shymkent":20,"turkestan":88,"kyzylorda":19},{"date":"2026-05-06","shymkent":20,"turkestan":141,"kyzylorda":16},{"date":"2026-05-07","shymkent":40,"turkestan":30,"kyzylorda":0},{"date":"2026-05-08","shymkent":25,"turkestan":54,"kyzylorda":17},{"date":"2026-05-12","shymkent":30,"turkestan":9,"kyzylorda":24},{"date":"2026-05-13","shymkent":22,"turkestan":0,"kyzylorda":22},{"date":"2026-05-14","shymkent":50,"turkestan":0,"kyzylorda":30},{"date":"2026-05-15","shymkent":48,"turkestan":0,"kyzylorda":42},{"date":"2026-05-18","shymkent":52,"turkestan":5,"kyzylorda":53},{"date":"2026-05-19","shymkent":39,"turkestan":0,"kyzylorda":59},{"date":"2026-05-20","shymkent":31,"turkestan":0,"kyzylorda":66},{"date":"2026-05-21","shymkent":45,"turkestan":0,"kyzylorda":54},{"date":"2026-05-22","shymkent":36,"turkestan":0,"kyzylorda":0},{"date":"2026-05-25","shymkent":0,"turkestan":0,"kyzylorda":2},{"date":"2026-05-26","shymkent":0,"turkestan":0,"kyzylorda":1},{"date":"2026-06-02","shymkent":35,"turkestan":4,"kyzylorda":0},{"date":"2026-06-03","shymkent":25,"turkestan":11,"kyzylorda":0},{"date":"2026-06-04","shymkent":33,"turkestan":11,"kyzylorda":0},{"date":"2026-06-05","shymkent":40,"turkestan":5,"kyzylorda":0},{"date":"2026-06-08","shymkent":0,"turkestan":9,"kyzylorda":0},{"date":"2026-06-09","shymkent":25,"turkestan":0,"kyzylorda":0},{"date":"2026-06-10","shymkent":30,"turkestan":10,"kyzylorda":0},{"date":"2026-06-11","shymkent":32,"turkestan":16,"kyzylorda":0},{"date":"2026-06-12","shymkent":29,"turkestan":19,"kyzylorda":0},{"date":"2026-06-15","shymkent":36,"turkestan":17,"kyzylorda":0},{"date":"2026-06-16","shymkent":35,"turkestan":10,"kyzylorda":0},{"date":"2026-06-17","shymkent":4,"turkestan":11,"kyzylorda":0},{"date":"2026-06-18","shymkent":42,"turkestan":10,"kyzylorda":0},{"date":"2026-06-19","shymkent":55,"turkestan":10,"kyzylorda":0},{"date":"2026-06-20","shymkent":57,"turkestan":0,"kyzylorda":0},{"date":"2026-06-21","shymkent":46,"turkestan":0,"kyzylorda":0},{"date":"2026-06-22","shymkent":39,"turkestan":16,"kyzylorda":2},{"date":"2026-06-23","shymkent":90,"turkestan":21,"kyzylorda":0},{"date":"2026-06-24","shymkent":32,"turkestan":7,"kyzylorda":0},{"date":"2026-06-25","shymkent":139,"turkestan":23,"kyzylorda":0},{"date":"2026-06-26","shymkent":91,"turkestan":24,"kyzylorda":0},{"date":"2026-06-27","shymkent":88,"turkestan":0,"kyzylorda":0},{"date":"2026-06-28","shymkent":1,"turkestan":0,"kyzylorda":0},{"date":"2026-06-29","shymkent":140,"turkestan":0,"kyzylorda":0},{"date":"2026-06-30","shymkent":162,"turkestan":0,"kyzylorda":0},{"date":"2026-07-01","shymkent":80,"turkestan":24,"kyzylorda":0},{"date":"2026-07-02","shymkent":61,"turkestan":16,"kyzylorda":0},{"date":"2026-07-03","shymkent":154,"turkestan":0,"kyzylorda":0},{"date":"2026-07-04","shymkent":187,"turkestan":0,"kyzylorda":0},{"date":"2026-07-05","shymkent":13,"turkestan":0,"kyzylorda":0},{"date":"2026-07-07","shymkent":165,"turkestan":0,"kyzylorda":0},{"date":"2026-07-08","shymkent":44,"turkestan":1,"kyzylorda":0},{"date":"2026-07-09","shymkent":61,"turkestan":0,"kyzylorda":0},{"date":"2026-07-10","shymkent":88,"turkestan":0,"kyzylorda":3},{"date":"2026-07-11","shymkent":38,"turkestan":0,"kyzylorda":7},{"date":"2026-07-12","shymkent":1,"turkestan":0,"kyzylorda":0},{"date":"2026-07-13","shymkent":90,"turkestan":0,"kyzylorda":0},{"date":"2026-07-14","shymkent":119,"turkestan":23,"kyzylorda":5},{"date":"2026-07-15","shymkent":175,"turkestan":51,"kyzylorda":33},{"date":"2026-07-16","shymkent":174,"turkestan":42,"kyzylorda":64},{"date":"2026-07-17","shymkent":213,"turkestan":20,"kyzylorda":77},{"date":"2026-07-18","shymkent":107,"turkestan":0,"kyzylorda":0},{"date":"2026-07-19","shymkent":9,"turkestan":0,"kyzylorda":0},{"date":"2026-07-20","shymkent":156,"turkestan":6,"kyzylorda":116},{"date":"2026-07-21","shymkent":218,"turkestan":29,"kyzylorda":133},{"date":"2026-07-22","shymkent":188,"turkestan":55,"kyzylorda":227},{"date":"2026-07-23","shymkent":158,"turkestan":78,"kyzylorda":243},{"date":"2026-07-24","shymkent":255,"turkestan":87,"kyzylorda":200},{"date":"2026-07-25","shymkent":140,"turkestan":0,"kyzylorda":0},{"date":"2026-07-27","shymkent":223,"turkestan":46,"kyzylorda":161},{"date":"2026-07-28","shymkent":169,"turkestan":104,"kyzylorda":184},{"date":"2026-07-29","shymkent":210,"turkestan":136,"kyzylorda":240},{"date":"2026-07-30","shymkent":155,"turkestan":119,"kyzylorda":174},{"date":"2026-07-31","shymkent":194,"turkestan":172,"kyzylorda":296}]''')

PROJECT_STATUS = {
    "systems": [
        "Приняты исходные коды ИС",
        "Описание бизнес-процессов — до 01.07.2026",
        "Разработка ТЗ — до 08.08.2026",
        "Ввод в опытную эксплуатацию — до 31.12.2026; промышленная эксплуатация — до 30.06.2027",
    ],
    "regulatory": [
        "Поправки в НПА подписаны Главой государства.",
        "Закон о переводе бытовых потребителей на интеллектуальные приборы учёта газа подписан Президентом Республики Казахстан.",
        "Передача данных от юридических лиц — до конца 2027 года за счёт потребителей.",
    ],
}

RU2EN = {"город Шымкент": "shymkent",
         "Туркестанская область": "turkestan",
         "Кызылординская область": "kyzylorda"}
EN2RU = {v: k for k, v in RU2EN.items()}
PF2EN = {"Шымкентский ПФ": "shymkent",
         "Туркестанский ПФ": "turkestan",
         "Кызылординский ПФ": "kyzylorda"}


# ──────────────────────────────────────────────────────────────
# ХЕЛПЕРЫ
# ──────────────────────────────────────────────────────────────

def die(msg):
    print(f"\n❌ ОШИБКА: {msg}\n", file=sys.stderr)
    sys.exit(1)


def clean_num(x):
    if x is None:
        return 0
    s = str(x).replace("\xa0", "").replace(" ", "").replace(",", "").strip()
    if s in ("", "nan", "None", "-"):
        return 0
    try:
        return int(float(s))
    except ValueError:
        return 0


def ru_date(d):
    return d.strftime("%d.%m.%Y")


def fmt(n):
    return f"{n:,}".replace(",", " ")


def working_days_between(a, b):
    if b < a:
        return 0
    return sum(1 for i in range((b - a).days + 1)
               if (a + timedelta(i)).weekday() < 5)


# ──────────────────────────────────────────────────────────────
# 1. Установки — Google Sheets (весь xlsx, все листы)
# ──────────────────────────────────────────────────────────────

def fetch_installs():
    print("→ Тяну отчёт по установке из Google Sheets…")
    try:
        resp = requests.get(INSTALL_XLSX_URL, timeout=60)
        resp.raise_for_status()
    except Exception as e:
        die(f"не скачался xlsx установок (шит открыт по ссылке?): {e}")

    if b"<html" in resp.content[:200].lower():
        die("Google вернул HTML вместо xlsx — шит закрыт (нужен доступ «по ссылке»).")

    import openpyxl
    wb = openpyxl.load_workbook(io.BytesIO(resp.content), data_only=True)

    if "Ввод данных" not in wb.sheetnames:
        die(f"в отчёте нет листа «Ввод данных». Листы: {wb.sheetnames}")

    ws = wb["Ввод данных"]

    # ── Сводка по ПФ (rows 7-9) ──
    installed = {}
    plan_by_reg = {}
    row_map = {7: "shymkent", 8: "turkestan", 9: "kyzylorda"}
    for r, en in row_map.items():
        name = ws.cell(r, 3).value
        if not name or PF2EN.get(str(name).strip()) != en:
            die(f"структура отчёта изменилась: строка {r} должна быть ПФ '{en}', "
                f"а там '{name}'. Проверь лист «Ввод данных».")
        installed[en] = clean_num(ws.cell(r, 6).value)   # Всего установлено
        plan_by_reg[en] = clean_num(ws.cell(r, 7).value)  # План 2026
    total_installed = sum(installed.values())

    if total_installed == 0:
        die("сумма установок = 0, проверь отчёт.")

    # Дата актуализации
    as_of = ws.cell(4, 5).value
    if isinstance(as_of, datetime):
        as_of = as_of.date()
    elif isinstance(as_of, str):
        try:
            as_of = datetime.strptime(as_of.strip(), "%d.%m.%Y").date()
        except ValueError:
            as_of = date.today()
    else:
        as_of = date.today()

    # ── Daily август: ШПФ и КПФ из своих листов (col1=дата, col5=за день),
    #    ТПФ из «Ввод данных» col H (лист ТПФ в шаблоне = копия ШПФ) ──
    def read_sheet_daily(sheet):
        out = {}
        for r in range(11, sheet.max_row + 1):
            d = sheet.cell(r, 1).value
            v = sheet.cell(r, 5).value
            if isinstance(d, datetime) and v and float(v) > 0:
                out[d.date()] = int(float(v))
        return out

    shpf = read_sheet_daily(wb["ШПФ"]) if "ШПФ" in wb.sheetnames else {}
    kpf = read_sheet_daily(wb["КПФ"]) if "КПФ" in wb.sheetnames else {}

    tpf = {}
    for r in range(17, 60):
        d = ws.cell(r, 6).value   # col F
        v = ws.cell(r, 8).value   # col H = ТПФ
        if isinstance(d, datetime) and v and float(v) > 0:
            tpf[d.date()] = int(float(v))

    aug_dates = sorted(set(shpf) | set(tpf) | set(kpf))
    daily_aug = []
    for d in aug_dates:
        daily_aug.append({
            "date": str(d),
            "shymkent": shpf.get(d, 0),
            "turkestan": tpf.get(d, 0),
            "kyzylorda": kpf.get(d, 0),
        })

    print(f"  ✓ установлено {fmt(total_installed)} "
          f"(ШПФ {fmt(installed['shymkent'])}, ТПФ {fmt(installed['turkestan'])}, "
          f"КПФ {fmt(installed['kyzylorda'])}); актуально на {ru_date(as_of)}")

    return {
        "installed": installed,
        "total_installed": total_installed,
        "plan_by_reg": plan_by_reg,
        "as_of": as_of,
        "daily_aug": daily_aug,
    }


# ──────────────────────────────────────────────────────────────
# 2. Склад — Google Sheets CSV
# ──────────────────────────────────────────────────────────────

def fetch_warehouse():
    print("→ Тяну склад из Google Sheets…")
    try:
        resp = requests.get(WAREHOUSE_CSV_URL, timeout=30)
        resp.raise_for_status()
    except Exception as e:
        die(f"не скачался CSV склада: {e}")

    if "<html" in resp.text[:200].lower():
        die("Google вернул HTML вместо CSV — склад-шит закрыт.")

    df = pd.read_csv(io.StringIO(resp.text), header=None)

    def cell(r, c):
        try:
            return str(df.iloc[r, c]).strip()
        except Exception:
            return ""

    checks = [
        (cell(1, 0).startswith("Всего отправлено"), 'row1[0] = "Всего отправлено"'),
        (cell(2, 0).startswith("Всего принято"),    'row2[0] = "Всего принято"'),
        (cell(0, 5).startswith("ТПФ"),              'row0[5] = "ТПФ"'),
        (cell(0, 6).startswith("ШПФ"),              'row0[6] = "ШПФ"'),
        (cell(0, 7).startswith("КПФ"),              'row0[7] = "КПФ"'),
    ]
    for ok, what in checks:
        if not ok:
            die(f"структура склада изменилась — ожидалось {what}. "
                f"Возможно вставлена строка/колонка. Скрипт остановлен.")

    # колонки: F=5 ТПФ, G=6 ШПФ, H=7 КПФ; блок «на складе сейчас»: M=12 N=13 O=14
    col = {"turkestan": 5, "shymkent": 6, "kyzylorda": 7}
    stock_col = {"turkestan": 12, "shymkent": 13, "kyzylorda": 14}
    field_col = {"turkestan": 12, "shymkent": 13, "kyzylorda": 14}  # row2 = отгружено в поле

    wh = {}
    for en, c in col.items():
        wh[en] = {
            "accepted":      clean_num(df.iloc[1, c]),                 # Всего отправлено
            "acceptedByDoc": clean_num(df.iloc[2, c]),                 # Всего принято
            "paid":          clean_num(df.iloc[3, c]),                 # Всего оплачено
            "stock":         clean_num(df.iloc[5, stock_col[en]]),     # на складе сейчас
            "toField":       clean_num(df.iloc[2, field_col[en]]),     # отгружено в поле
        }

    tot = {k: sum(v[k] for v in wh.values())
           for k in ("accepted", "acceptedByDoc", "paid", "stock", "toField")}
    print(f"  ✓ склад: отправлено {fmt(tot['accepted'])}, "
          f"принято {fmt(tot['acceptedByDoc'])}, оплачено {fmt(tot['paid'])}")
    wh["_total"] = tot
    return wh


# ──────────────────────────────────────────────────────────────
# 3. Активировано + e-Qural total — PostgreSQL
# ──────────────────────────────────────────────────────────────

def fetch_platform():
    """Активировано — через e-Qural API GetActivatedCount.
       К установке — из БД MMS (device_status='2')."""
    # ── 1. Активировано (e-Qural API) — ВРЕМЕННО ОТКЛЮЧЕНО ──
    # Сеть: equral.ktga.kz (10.20.14.10) недоступен с VPN-сегмента,
    # ждём открытие доступа от сетевого инженера QGA. Пока activated = 0.
    activated = {"shymkent": 0, "turkestan": 0, "kyzylorda": 0}
    act_total = 0

    # ── 2. К установке (БД MMS, device_status='2') ──
    print("→ Подключаюсь к БД MMS…")
    try:
        import pymysql
    except ModuleNotFoundError:
        die("нет pymysql — pip install --break-system-packages pymysql")
    try:
        conn = pymysql.connect(connect_timeout=15, **DB_MMS)
    except Exception as e:
        die(f"БД MMS недоступна (VPN?): {e}")

    to_install = {"shymkent": 0, "turkestan": 0, "kyzylorda": 0}
    try:
        cur = conn.cursor()
        cur.execute(
            "SELECT region_code, COUNT(*) FROM dl_device "
            "WHERE device_status = %s GROUP BY region_code",
            (MMS_STATUS["toInstall"],))
        for region_code, cnt in cur.fetchall():
            en = MMS_REGION.get(str(region_code))
            if en:
                to_install[en] = int(cnt)
    except Exception as e:
        die(f"SQL MMS упал: {e}")
    finally:
        conn.close()
    to_install_total = sum(to_install.values())

    print(f"  ✓ активировано {fmt(act_total)}, к установке {fmt(to_install_total)}")
    # online пока не считаем отдельно (нет источника) — 0
    online = {"shymkent": 0, "turkestan": 0, "kyzylorda": 0}
    return {"activated": activated, "online": online,
            "activated_total": act_total,
            "toInstall": to_install, "toInstall_total": to_install_total}


# ──────────────────────────────────────────────────────────────
# 4. Сборка DATA
# ──────────────────────────────────────────────────────────────

def build_data(inst, wh, plat):
    print("→ Считаю KPI…")

    installed = inst["installed"]
    total = inst["total_installed"]
    as_of = inst["as_of"]
    activated = plat["activated"]
    online = plat["online"]
    online_total = sum(online.values())
    wt = wh["_total"]

    # ── Региональный срез ──
    regional = []
    for en in ("kyzylorda", "turkestan", "shymkent"):
        ru = EN2RU[en]
        inst_n = installed[en]
        w = wh[en]
        rp = REGIONAL_PLANS[ru]
        stock_remain = w["acceptedByDoc"] - inst_n           # ТЗ п.3: Принято − Факт всего
        fill_base = w["accepted"] - w["toField"]             # ТЗ п.4
        fill_pct = round(fill_base / CAPACITY[en] * 100, 2) if CAPACITY[en] else 0.0
        regional.append({
            "region": ru,
            "installed": inst_n,
            "activated": activated[en],
            "online": online[en],
            "offline": inst_n - online[en],
            "availability": round(online[en] / inst_n * 100, 2) if inst_n else 0.0,
            "accepted": w["accepted"],
            "acceptedByDoc": w["acceptedByDoc"],
            "paidAmount": w["paid"],
            "stockRemaining": stock_remain,
            "warehouseFillPct": fill_pct,
            "plan2026": rp["plan2026"],
            "plan2027": rp["plan2027"],
            "phasePlan": rp["phasePlan"],
            "progress2026": round(inst_n / rp["plan2026"] * 100, 2) if rp["plan2026"] else 0.0,
        })

    # ── Daily: история (апр-июл) + август из отчёта ──
    daily = []
    for h in HISTORICAL_DAILY:
        d = date.fromisoformat(h["date"])
        regions = {}
        for en in ("shymkent", "turkestan", "kyzylorda"):
            if h[en] > 0:
                regions[EN2RU[en]] = {"installed": h[en], "online": h[en], "offline": 0}
        if regions:
            daily.append({"date": h["date"], "label": ru_date(d), "regions": regions})

    for a in inst["daily_aug"]:
        d = date.fromisoformat(a["date"])
        regions = {}
        for en in ("shymkent", "turkestan", "kyzylorda"):
            if a[en] > 0:
                regions[EN2RU[en]] = {"installed": a[en], "online": a[en], "offline": 0}
        if regions:
            daily.append({"date": a["date"], "label": ru_date(d), "regions": regions})

    # ── Месячный план vs факт ──
    monthly_facts = {}
    for entry in daily:
        mo = date.fromisoformat(entry["date"]).month
        day_total = sum(v["installed"] for v in entry["regions"].values())
        monthly_facts[mo] = monthly_facts.get(mo, 0) + day_total

    monthly_plan = []
    cum_plan = cum_fact = 0
    for mo in range(1, 13):
        mp = MONTHLY_PLAN.get(mo, 0)
        cum_plan += mp
        mf = monthly_facts.get(mo, 0)
        cum_fact += mf
        monthly_plan.append({
            "date": f"2026-{mo:02d}-01", "label": f"{mo:02d}.2026",
            "year": 2026, "month": mo, "phase1": mp, "phase2": 0, "total": mp,
            "cumulativePlan": cum_plan, "fact": mf, "cumulativeFact": cum_fact,
        })

    # ── Темпы / прогноз ──
    today = date.today()
    wd_left = working_days_between(today, YEAR_END)
    cal_left = (YEAR_END - today).days + 1
    remaining = PLAN_2026 - total
    rec_pace = round(remaining / wd_left, 1) if wd_left > 0 else 0.0

    # темп по последним 5 дням с данными
    recent = [sum(v["installed"] for v in e["regions"].values()) for e in daily[-5:]]
    avg_pace = round(sum(recent) / len(recent), 1) if recent else 0.0
    if avg_pace > 0:
        forecast_date_str = ru_date(today + timedelta(days=remaining / avg_pace))
    else:
        forecast_date_str = "н/д"
    forecast_volume = total + round(avg_pace * cal_left)

    lag = PLAN_TO_AUG - total
    exec_pct = round(total / PLAN_TO_AUG * 100, 2)

    # ── Статусные тексты ──
    ps = json.loads(json.dumps(PROJECT_STATUS))
    ps["meters"] = [
        f"Установлено {fmt(total)} smart-счётчиков (активировано {fmt(plat['activated_total'])})",
        f"Отправлено с завода {fmt(wt['accepted'])} шт.; принято по документам — "
        f"{fmt(wt['acceptedByDoc'])} шт.; остаток — {fmt(wt['stock'])} шт.",
        "Договоры на монтаж заключены 16.06.2026 г.",
        "Август 2026: новый расчётный период.",
    ]
    ps["open"] = [
        f"Отставание от планового графика — {fmt(lag)} счётчиков на {ru_date(as_of)}",
        "Требуется ускорение монтажа и завершение закупок услуг связи",
        "Отчёт перед СД направлен в ДИТ АО «НК «QazaqGaz»",
    ]

    summary = {
        "generatedAt": ru_date(today),
        "installed": total,
        "activated": plat["activated_total"],
        "online": online_total,
        "offline": total - online_total,
        "availability": round(online_total / total * 100, 2) if total else 0,
        "accepted": wt["accepted"],
        "acceptedByDoc": wt["acceptedByDoc"],
        "totalPaidAmount": wt["paid"],
        "stockRemaining": wt["acceptedByDoc"] - total,
        "toInstallTotal": plat["toInstall_total"],
        "plan2026Presentation": 199000,
        "phase1PlanPresentation": PHASE1_PLAN,
        "totalProgramPlan": TOTAL_PROGRAM_PLAN,
        "progress2026": round(total / PLAN_2026 * 100, 2),
        "progressPhase1": round(total / PHASE1_PLAN * 100, 2),
        "minDate": str(PROJECT_START),
        "maxDate": str(as_of),
        "delayDays": (as_of - PROJECT_START).days,
        "planToCurrentMonth": PLAN_TO_AUG,
        "lagToCurrentMonth": lag,
        "executionToCurrentMonth": exec_pct,
        "asOfMonthLabel": as_of.strftime("%m.%Y"),
        "annualPlan2026Monthly": PLAN_2026,
        "remaining2026Plan": remaining,
        "recommendedPace": rec_pace,
        "avgPace10": avg_pace,
        "forecastDateStr": forecast_date_str,
        "forecastVolume2026": forecast_volume,
        "daysLeftToYearEnd": cal_left,
        "workingDaysLeft": wd_left,
    }

    print(f"  ✓ installed={fmt(total)} activated={fmt(plat['activated_total'])} "
          f"lag={fmt(lag)} pace={avg_pace} forecast={forecast_date_str}")

    return {
        "summary": summary,
        "regional": regional,
        "daily": daily,
        "schedule": ANNUAL_SCHEDULE,
        "monthlyPlan": monthly_plan,
        "annualPlanFromMonthly": ANNUAL_SCHEDULE,
        "projectStatus": ps,
    }


# ──────────────────────────────────────────────────────────────
# 5. Сборка index.html из шаблона
# ──────────────────────────────────────────────────────────────

def render(data, repo_dir):
    """Standalone-дизайн читает window.__DASH_DATA из
       <script src="data:text/javascript;base64,...">. Пересобираем этот блок."""
    import os, base64
    tpl_path = os.path.join(os.path.expanduser(repo_dir), TEMPLATE_FILE)
    if not os.path.exists(tpl_path):
        die(f"нет шаблона {TEMPLATE_FILE} в {repo_dir}")
    with open(tpl_path, encoding="utf-8") as f:
        html = f.read()

    payload = json.dumps(data, ensure_ascii=False)
    js = f"window.__DASH_DATA = {payload};"
    b64 = base64.b64encode(js.encode("utf-8")).decode("ascii")

    new_src = f'src="data:text/javascript;base64,{b64}"'
    html2, n = re.subn(
        r'src="data:text/javascript;base64,[A-Za-z0-9+/=]+"',
        new_src, html, count=1)
    if n != 1:
        die("не найден блок window.__DASH_DATA (data:base64) в шаблоне.")

    # sanity: декодируем обратно и парсим
    m = re.search(r'src="data:text/javascript;base64,([A-Za-z0-9+/=]+)"', html2)
    dec = base64.b64decode(m.group(1)).decode("utf-8")
    jm = re.search(r'window\.__DASH_DATA\s*=\s*(\{.*\});?$', dec, re.DOTALL)
    try:
        json.loads(jm.group(1).rstrip(";"))
    except Exception as e:
        die(f"DASH_DATA не парсится после сборки: {e}")

    out_path = os.path.join(os.path.expanduser(repo_dir), INDEX_FILE)
    with open(out_path, "w", encoding="utf-8") as f:
        f.write(html2)
    print(f"  ✓ {INDEX_FILE} собран ({len(html2):,} байт)".replace(",", " "))
    return out_path


# ──────────────────────────────────────────────────────────────
# 6. Git push
# ──────────────────────────────────────────────────────────────

def git_push(repo_dir):
    import os
    d = os.path.expanduser(repo_dir)
    print("→ git push…")

    def run(*args):
        return subprocess.run(args, cwd=d, capture_output=True, text=True)

    if run("git", "add", INDEX_FILE).returncode != 0:
        die("git add упал")
    if not run("git", "status", "--porcelain").stdout.strip():
        print("  ⚠ нет изменений — пропускаю commit.")
        return
    msg = f"Auto update dashboard {datetime.now():%d.%m.%Y %H:%M}"
    c = run("git", "commit", "-m", msg)
    if c.returncode != 0:
        die(f"git commit упал: {c.stderr.strip() or c.stdout.strip()}")
    p = run("git", "push")
    if p.returncode != 0:
        die(f"git push упал: {p.stderr.strip()}")
    print("  ✓ запушено → Cloudflare передеплоит → MMS iframe обновится")


# ──────────────────────────────────────────────────────────────
# MAIN
# ──────────────────────────────────────────────────────────────

def main():
    no_git = ("--dry-run" in sys.argv) or ("--no-git" in sys.argv)

    print("=" * 55)
    print("  Aqyl Energy — обновление CEO-дашборда (v2)")
    print("=" * 55)

    inst = fetch_installs()
    wh = fetch_warehouse()
    plat = fetch_platform()
    data = build_data(inst, wh, plat)
    render(data, REPO_DIR)

    if no_git:
        print("\n(--dry-run) git пропущен. index.html собран локально.")
    else:
        git_push(REPO_DIR)

    print("\n✅ Готово.\n")


if __name__ == "__main__":
    main()
